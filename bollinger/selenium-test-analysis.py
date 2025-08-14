# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
import argparse
import sys

# Noms des feuilles d'analyse (constants)
SHEET_ALL = "All_Results"
SHEET_BEST_PER_SYMBOL = "Analysis_Best_Per_Symbol"
SHEET_BEST_GLOBAL = "Analysis_Best_Global"
SHEET_COMBO_COUNTS = "Analysis_Combo_Counts"
SHEET_ALL_SCORED = "Analysis_All_Scored"
SHEET_GLOBAL_VS_CUSTOM = "Analysis_Global_vs_Custom"

# === CONFIG: Hard filters & scoring weights (tweak here) ===
MIN_TRADES      = 30        # discard rows with fewer trades
MAX_DRAWDOWN    = 15.0      # discard rows with drawdown (%) above this
MIN_PROFIT_FACT = 1.2       # discard rows with PF below this
MIN_WIN_RATE    = 25.0      # discard rows with win rate (%) below this

SCORE_WEIGHTS = {
    "profit": 1.0,   # weight for Net Profit Clean (absolute)
    "wr": 0.5,       # weight for Win Rate (0-100)
    "pf": 3.0,       # weight for Profit Factor
    "dd": 1.0        # penalty weight for drawdown (higher drawdown reduces score)
}

def load_all_results(path: str, sheet_all: str) -> pd.DataFrame:
    wb = load_workbook(path)
    if sheet_all in wb.sheetnames:
        return pd.read_excel(path, sheet_name=sheet_all, engine="openpyxl")

    # Fallback: merge all *_Results sheets
    dfs = []
    for s in wb.sheetnames:
        if s.endswith("_Results"):
            df_s = pd.read_excel(path, sheet_name=s, engine="openpyxl")
            df_s["Symbol"] = s.replace("_Results", "")
            dfs.append(df_s)
    if not dfs:
        raise RuntimeError("No All_Results or *_Results sheets found.")
    return pd.concat(dfs, ignore_index=True)

def clean_numeric(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["Net Profit Clean"] = pd.to_numeric(out["Net Profit"].replace(r"[\$,]", "", regex=True), errors="coerce")
    out["Win Rate"]        = pd.to_numeric(out["Win Rate"].replace(r"[\%,]", "", regex=True), errors="coerce")
    out["drawdown"]        = pd.to_numeric(out["drawdown"].replace(r"[\%,]", "", regex=True), errors="coerce")
    out["Profit Factor"]   = pd.to_numeric(out["Profit Factor"].replace(r"[x,]", "", regex=True), errors="coerce")
    out["Total Trades"]    = pd.to_numeric(out["Total Trades"].replace(r"[^\d]", "", regex=True), errors="coerce")
    # Optional: ensure parameter columns numeric
    if "ATR Multiplier" in out.columns:
        out["ATR Multiplier"] = pd.to_numeric(out["ATR Multiplier"], errors="coerce")
    if "RR" in out.columns:
        out["RR"] = pd.to_numeric(out["RR"], errors="coerce")
    if "Vol Multiplier" in out.columns:
        out["Vol Multiplier"] = pd.to_numeric(out["Vol Multiplier"], errors="coerce")
    return out


# === Helper: Apply hard filters based on CONFIG ===
def apply_hard_filters(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    # Ensure numeric
    for col in ["Total Trades", "drawdown", "Profit Factor", "Win Rate"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    # Build mask with available columns
    mask = pd.Series(True, index=out.index)
    if "Total Trades" in out.columns:
        mask &= out["Total Trades"] >= MIN_TRADES
    if "drawdown" in out.columns:
        mask &= out["drawdown"] <= MAX_DRAWDOWN
    if "Profit Factor" in out.columns:
        mask &= out["Profit Factor"] >= MIN_PROFIT_FACT
    if "Win Rate" in out.columns:
        mask &= out["Win Rate"] >= MIN_WIN_RATE
    return out[mask].reset_index(drop=True)

def compute_score(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    # Fill NaNs
    profit = out["Net Profit Clean"].fillna(0)
    wr     = out["Win Rate"].fillna(0) / 100.0  # 0..1
    pf     = out["Profit Factor"].fillna(0)
    dd     = out["drawdown"].fillna(0)

    w = SCORE_WEIGHTS
    # Score = +profit*w1 + wr*100*w2 + pf*w3 - dd*w4
    out["Score"] = (
        profit * w.get("profit", 1.0)
        + (wr * 100.0) * w.get("wr", 0.5)
        + pf * w.get("pf", 3.0)
        - dd * w.get("dd", 1.0)
    )
    return out

def best_per_symbol(df_scored: pd.DataFrame) -> pd.DataFrame:
    # Keep only fully scored rows
    ok = df_scored.dropna(subset=["Symbol", "Score"])
    best_idx = ok.groupby("Symbol")["Score"].idxmax()
    return ok.loc[best_idx].sort_values("Score", ascending=False).reset_index(drop=True)

def best_global_avg(df_scored: pd.DataFrame) -> pd.DataFrame:
    # Rank parameter sets by average Score across symbols
    # Keep only rows with parameters present
    needed = ["ATR Multiplier", "RR", "Vol Multiplier"]
    have = [c for c in needed if c in df_scored.columns]
    if not have:
        raise RuntimeError("No parameter columns found to compute global averages.")
    grouped = (
        df_scored.dropna(subset=["Score"] + have)
                 .groupby(have)
                 .agg({
                     "Score": "mean",
                     "Net Profit Clean": "mean",
                     "drawdown": "mean",
                     "Win Rate": "mean",
                     "Profit Factor": "mean",
                     "Total Trades": "mean",
                 })
                 .sort_values("Score", ascending=False)
                 .reset_index()
    )
    return grouped

def combo_counts_among_winners(df_best_per_symbol: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in ["ATR Multiplier", "RR", "Vol Multiplier"] if c in df_best_per_symbol.columns]
    if not cols:
        return pd.DataFrame(columns=["Count"])
    # Robustesse : nombre de fois où chaque combinaison est dans le top, et nombre de symboles différents
    grouped = df_best_per_symbol.groupby(cols)
    counts = grouped.size().reset_index(name="Count")
    # Ajoute le nombre de symboles distincts pour chaque combinaison
    if "Symbol" in df_best_per_symbol.columns:
        counts["NumSymbols"] = grouped["Symbol"].nunique().values
    counts = counts.sort_values(["NumSymbols", "Count"], ascending=[False, False]).reset_index(drop=True)
    return counts

def write_analysis(path: str,
                   all_scored: pd.DataFrame,
                   best_per_sym: pd.DataFrame,
                   best_global: pd.DataFrame,
                   combo_counts: pd.DataFrame,
                   global_vs_custom: pd.DataFrame,
                   sheet_names: dict):
    """Écrit l'analyse dans le fichier Excel avec styling."""
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        all_scored.to_excel(writer, index=False, sheet_name=sheet_names['all_scored'])
        best_per_sym.to_excel(writer, index=False, sheet_name=sheet_names['best_per_symbol'])
        best_global.to_excel(writer, index=False, sheet_name=sheet_names['best_global'])
        combo_counts.to_excel(writer, index=False, sheet_name=sheet_names['combo_counts'])
        global_vs_custom.to_excel(writer, index=False, sheet_name=sheet_names['global_vs_custom'])

    # Styling
    wb = load_workbook(path)

    def style_sheet(ws_name: str, highlight_top5: bool = False, add_scales: bool = False):
        if ws_name not in wb.sheetnames:
            return
        ws = wb[ws_name]

        # Header style
        header_fill = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Highlight top 5
        if highlight_top5:
            green_fill = PatternFill(start_color="FFDFF0D8", end_color="FFDFF0D8", fill_type="solid")
            top_n = min(5, ws.max_row - 1)
            for r in range(2, 2 + top_n):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = green_fill

        # Conditional color scales
        if add_scales:
            headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
            from openpyxl.utils import get_column_letter
            start_row, end_row = 2, ws.max_row

            def add_scale(col_name: str, good_high=True):
                if col_name not in headers or end_row < start_row:
                    return
                col = headers[col_name]
                rng = f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}"
                if good_high:
                    ws.conditional_formatting.add(
                        rng,
                        ColorScaleRule(
                            start_type="min", start_color="FFF4CCCC",   # red
                            mid_type="percentile", mid_value=50, mid_color="FFFFFFCC",  # yellow
                            end_type="max", end_color="FFDFF0D8"        # green
                        )
                    )
                else:
                    ws.conditional_formatting.add(
                        rng,
                        ColorScaleRule(
                            start_type="min", start_color="FFDFF0D8",   # green (low is good)
                            mid_type="percentile", mid_value=50, mid_color="FFFFFFCC",
                            end_type="max", end_color="FFF4CCCC"        # red (high is bad)
                        )
                    )

            for col in ["Score", "Net Profit Clean", "Win Rate", "Profit Factor"]:
                add_scale(col, good_high=True)
            add_scale("drawdown", good_high=False)

        # Autosize columns
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    style_sheet(sheet_names['best_global'], highlight_top5=True, add_scales=True)
    style_sheet(sheet_names['best_per_symbol'], highlight_top5=False, add_scales=True)
    style_sheet(sheet_names['all_scored'], highlight_top5=False, add_scales=True)
    style_sheet(sheet_names['combo_counts'], highlight_top5=False, add_scales=False)
    style_sheet(sheet_names['global_vs_custom'], highlight_top5=False, add_scales=True)

    wb.save(path)


def compare_global_vs_custom(df_scored: pd.DataFrame, best_per_sym: pd.DataFrame, best_global: pd.DataFrame) -> pd.DataFrame:
    """Compare l'approche globale vs personnalisée par paire"""
    if best_global.empty or best_per_sym.empty:
        return pd.DataFrame()
    
    # Au lieu du meilleur absolu, trouvons le paramètre testé sur le plus de symboles
    param_coverage = []
    for _, row in best_global.iterrows():
        atr, rr, vol = row["ATR Multiplier"], row["RR"], row["Vol Multiplier"]
        symbols_tested = len(df_scored[
            (df_scored["ATR Multiplier"] == atr) &
            (df_scored["RR"] == rr) &
            (df_scored["Vol Multiplier"] == vol)
        ]["Symbol"].unique())
        param_coverage.append({
            "ATR": atr, "RR": rr, "Vol": vol,
            "Score": row["Score"], "Symbols_Tested": symbols_tested
        })
    
    coverage_df = pd.DataFrame(param_coverage)
    
    # Trouver la couverture maximale
    max_coverage = coverage_df["Symbols_Tested"].max()
    
    # Parmi ceux avec la couverture maximale, prendre celui avec le meilleur score
    max_coverage_params = coverage_df[coverage_df["Symbols_Tested"] == max_coverage]
    best_coverage = max_coverage_params.loc[max_coverage_params["Score"].idxmax()]
    
    global_atr = best_coverage["ATR"]
    global_rr = best_coverage["RR"] 
    global_vol = best_coverage["Vol"]
    
    print(f"[INFO] Paramètres globaux sélectionnés (meilleure couverture): ATR={global_atr}, RR={global_rr}, Vol={global_vol}")
    print(f"[INFO] Testés sur {best_coverage['Symbols_Tested']} symboles (score moyen: {best_coverage['Score']:.2f})")
    print(f"[INFO] Nombre de symboles à analyser: {len(best_per_sym)}")
    
    # Pour chaque symbole, comparer le score avec paramètre global vs optimal
    comparison = []
    
    for _, row in best_per_sym.iterrows():
        symbol = row["Symbol"]
        custom_score = row["Score"]
        
        # Trouver le score avec les paramètres globaux pour ce symbole
        global_mask = (
            (df_scored["Symbol"] == symbol) &
            (df_scored["ATR Multiplier"] == global_atr) &
            (df_scored["RR"] == global_rr) &
            (df_scored["Vol Multiplier"] == global_vol)
        )
        
        global_result = df_scored[global_mask]
        if not global_result.empty:
            global_score = global_result.iloc[0]["Score"]
            regret = custom_score - global_score  # Différence de performance
            regret_pct = (regret / custom_score * 100) if custom_score != 0 else 0
            
            comparison.append({
                "Symbol": symbol,
                "Custom_Score": custom_score,
                "Global_Score": global_score,
                "Regret": regret,
                "Regret_Pct": regret_pct,
                "Custom_ATR": row.get("ATR Multiplier"),
                "Custom_RR": row.get("RR"),
                "Custom_Vol": row.get("Vol Multiplier"),
                "Global_ATR": global_atr,
                "Global_RR": global_rr,
                "Global_Vol": global_vol
            })
        else:
            # Prendre le meilleur paramètre disponible pour ce symbole qui ressemble le plus au global
            symbol_data = df_scored[df_scored["Symbol"] == symbol]
            if not symbol_data.empty:
                # Calculer la "distance" euclidienne aux paramètres globaux
                symbol_data = symbol_data.copy()
                symbol_data["distance"] = (
                    (symbol_data["ATR Multiplier"] - global_atr)**2 +
                    (symbol_data["RR"] - global_rr)**2 +
                    (symbol_data["Vol Multiplier"] - global_vol)**2
                )**0.5
                
                closest_params = symbol_data.loc[symbol_data["distance"].idxmin()]
                fallback_score = closest_params["Score"]
                
                regret = custom_score - fallback_score
                regret_pct = (regret / custom_score * 100) if custom_score != 0 else 0
                
                comparison.append({
                    "Symbol": f"{symbol} (fallback)",
                    "Custom_Score": custom_score,
                    "Global_Score": fallback_score,
                    "Regret": regret,
                    "Regret_Pct": regret_pct,
                    "Custom_ATR": row.get("ATR Multiplier"),
                    "Custom_RR": row.get("RR"),
                    "Custom_Vol": row.get("Vol Multiplier"),
                    "Global_ATR": f"{closest_params['ATR Multiplier']} (closest)",
                    "Global_RR": f"{closest_params['RR']} (closest)",
                    "Global_Vol": f"{closest_params['Vol Multiplier']} (closest)"
                })
            else:
                # Si vraiment aucune donnée pour ce symbole
                comparison.append({
                    "Symbol": symbol,
                    "Custom_Score": custom_score,
                    "Global_Score": "N/A",
                    "Regret": "N/A",
                    "Regret_Pct": "N/A",
                    "Custom_ATR": row.get("ATR Multiplier"),
                    "Custom_RR": row.get("RR"),
                    "Custom_Vol": row.get("Vol Multiplier"),
                    "Global_ATR": global_atr,
                    "Global_RR": global_rr,
                    "Global_Vol": global_vol
                })
    
    comparison_df = pd.DataFrame(comparison)
    
    if not comparison_df.empty:
        # Filtrer les lignes avec des données valides pour les statistiques
        valid_rows = comparison_df[
            (comparison_df["Regret"] != "N/A") & 
            (comparison_df["Global_Score"] != "N/A")
        ].copy()
        
        if not valid_rows.empty:
            # Statistiques de synthèse
            avg_regret = valid_rows["Regret"].mean()
            avg_regret_pct = valid_rows["Regret_Pct"].mean()
            max_regret = valid_rows["Regret"].max()
            min_regret = valid_rows["Regret"].min()
            symbols_count = len(valid_rows)
            
            # Ajouter une ligne de synthèse
            summary = {
                "Symbol": f"SUMMARY ({symbols_count} symbols)",
                "Custom_Score": valid_rows["Custom_Score"].mean(),
                "Global_Score": valid_rows["Global_Score"].mean(),
                "Regret": avg_regret,
                "Regret_Pct": avg_regret_pct,
                "Custom_ATR": f"AVG_REGRET: {avg_regret:.2f}",
                "Custom_RR": f"MAX_REGRET: {max_regret:.2f}",
                "Custom_Vol": f"MIN_REGRET: {min_regret:.2f}",
                "Global_ATR": global_atr,
                "Global_RR": global_rr,
                "Global_Vol": global_vol
            }
            
            comparison_df = pd.concat([comparison_df, pd.DataFrame([summary])], ignore_index=True)
        
        # Tri sécurisé : convertir "N/A" en NaN pour le tri, puis les placer à la fin
        comparison_df["Regret_Sort"] = pd.to_numeric(comparison_df["Regret"], errors='coerce')
        comparison_df = comparison_df.sort_values("Regret_Sort", ascending=False, na_position='last').reset_index(drop=True)
        comparison_df = comparison_df.drop(columns="Regret_Sort")
    
    return comparison_df


# === Main: Apply hard filters before scoring, then analyze and write ===
def main():
    # Analyse des arguments de ligne de commande
    parser = argparse.ArgumentParser(description="Analyse des résultats de backtest TradingView")
    parser.add_argument('--level', choices=['COARSE', 'FINE', 'FULL'], default='FINE',
                       help='Niveau de test à analyser: COARSE, FINE, ou FULL (défaut: FINE)')
    args = parser.parse_args()
    
    # Configuration du fichier selon le niveau
    file_path = f"tradingview_backtest_results_{args.level.lower()}.xlsx"
    
    # Configuration des noms de feuilles
    sheet_names = {
        'all': SHEET_ALL,
        'all_scored': SHEET_ALL_SCORED,
        'best_per_symbol': SHEET_BEST_PER_SYMBOL,
        'best_global': SHEET_BEST_GLOBAL,
        'combo_counts': SHEET_COMBO_COUNTS,
        'global_vs_custom': SHEET_GLOBAL_VS_CUSTOM
    }
    
    print(f"🔍 ANALYSE DU NIVEAU {args.level}")
    print(f"📁 Fichier: {file_path}")
    
    # Vérifier que le fichier existe
    import os
    if not os.path.exists(file_path):
        print(f"❌ Erreur: Le fichier {file_path} n'existe pas.")
        print(f"💡 Assurez-vous d'avoir exécuté les tests pour le niveau {args.level}:")
        print(f"   python3 test-selenium-single-thread.py --level {args.level}")
        sys.exit(1)
    
    df_raw = load_all_results(file_path, sheet_names['all'])
    df_clean = clean_numeric(df_raw)

    # Apply hard filters BEFORE scoring
    df_filtered = apply_hard_filters(df_clean)

    # Compute score on filtered data
    df_scored = compute_score(df_filtered)

    # Derive analytics
    best_per_sym = best_per_symbol(df_scored)
    best_global = best_global_avg(df_scored)
    combo_counts = combo_counts_among_winners(best_per_sym)
    global_vs_custom = compare_global_vs_custom(df_scored, best_per_sym, best_global)

    # Write & style
    write_analysis(file_path, df_scored, best_per_sym, best_global, combo_counts, global_vs_custom, sheet_names)
    print(f"\n✅ Analyse {args.level} terminée. Feuilles créées dans {file_path}:")
    print(f" - {sheet_names['all_scored']}")
    print(f" - {sheet_names['best_per_symbol']}")
    print(f" - {sheet_names['best_global']}")
    print(f" - {sheet_names['combo_counts']}")
    print(f" - {sheet_names['global_vs_custom']}")
    
    # Statistiques du niveau analysé
    total_rows = len(df_raw)
    filtered_rows = len(df_filtered)
    symbols_count = df_raw['Symbol'].nunique() if 'Symbol' in df_raw.columns else 0
    
    print(f"\n📊 STATISTIQUES NIVEAU {args.level}:")
    print(f"   📈 {total_rows:,} résultats bruts")
    print(f"   ✅ {filtered_rows:,} résultats après filtres")
    print(f"   🎯 {symbols_count} symboles analysés")
    print(f"   🏆 Taux de réussite: {(filtered_rows/total_rows*100):.1f}%" if total_rows > 0 else "   🏆 Taux de réussite: N/A")
    
    # Afficher un résumé de la comparaison
    if not global_vs_custom.empty:
        summary_row = global_vs_custom[global_vs_custom["Symbol"].str.contains("SUMMARY", na=False)]
        if not summary_row.empty:
            avg_regret = summary_row.iloc[0]["Regret_Pct"]
            symbols_analyzed = len(global_vs_custom) - 1  # -1 pour exclure la ligne SUMMARY
            print(f"\n=== RECOMMANDATION ({symbols_analyzed} symboles analysés) ===")
            if abs(avg_regret) < 5:
                print(f"Regret moyen: {avg_regret:.1f}% → Utilisez les PARAMÈTRES GLOBAUX (plus robuste)")
            elif avg_regret > 10:
                print(f"Regret moyen: {avg_regret:.1f}% → Utilisez les PARAMÈTRES CUSTOM (gain significatif)")
            else:
                print(f"Regret moyen: {avg_regret:.1f}% → Compromis à évaluer selon votre tolérance au risque")
        else:
            print(f"\n=== INFO ===")
            print(f"Analyse de {len(global_vs_custom)} symboles effectuée, voir feuille {sheet_names['global_vs_custom']}")

if __name__ == "__main__":
    main()