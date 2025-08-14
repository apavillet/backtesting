# -*- coding: utf-8 -*-

import time
import pandas as pd
import os
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys


from selenium.webdriver import ActionChains

# --- Timing / profiling helpers ---
import statistics
from contextlib import contextmanager
from collections import defaultdict
from collections import deque
import sys
import argparse

TIMING_CSV = "tv_timings.csv"
timing_stats = defaultdict(list)       # label -> list of durations (seconds)
combo_timings_rows = []                # detailed per-combination rows for CSV

# --- ETA / progress helpers ---
combo_time_window = deque(maxlen=100)
DEFAULT_COMBO_SEC = 5.0  # fallback when we don't have timing yet

def _avg_combo_seconds():
    if combo_time_window:
        return sum(combo_time_window) / len(combo_time_window)
    if timing_stats.get("combo_total"):
        vals = timing_stats["combo_total"]
        if vals:
            return sum(vals) / len(vals)
    return DEFAULT_COMBO_SEC

def _fmt_eta(seconds: float):
    if seconds is None or seconds != seconds or seconds < 0:
        return "--:--"
    m, s = divmod(int(round(seconds)), 60)
    if m >= 60:
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"

def _print_progress(symbol: str, done: int, total: int, gdone: int, gtotal: int):
    avg_sec = _avg_combo_seconds()
    eta_symbol = (total - done) * avg_sec
    eta_all = (gtotal - gdone) * avg_sec
    msg = (
        f"\r[{symbol}] {done}/{total}  |  ETA sym { _fmt_eta(eta_symbol) }"
        f"  |  ETA all { _fmt_eta(eta_all) }  | avg/combo ~ {avg_sec:.2f}s"
    )
    try:
        sys.stdout.write(msg)
        sys.stdout.flush()
    except Exception:
        print(msg)
DEBUG = False  # Set to False to disable debug prints
@contextmanager
def timed(label: str):
    """Context manager to time code sections and collect aggregated stats."""
    t0 = time.perf_counter()
    try:
        yield
    finally:
        dt = time.perf_counter() - t0
        timing_stats[label].append(dt)
        if DEBUG:
            print(f"[Timing] {label}: {dt:.3f} seconds")


def _append_combo_timing_row(symbol, atr, rr, vol_mult, **parts):
    row = {
        "symbol": symbol,
        "atr": atr,
        "rr": rr,
        "vol_mult": vol_mult,
    }
    row.update(parts)
    combo_timings_rows.append(row)


def dump_timing_summary():
    print("\n==== TIMING SUMMARY (seconds) ====")
    for label, values in timing_stats.items():
        if not values:
            continue
        print(f"{label:22s}  n={len(values):4d}  avg={statistics.mean(values):6.3f}  p90={statistics.quantiles(values, n=10)[8]:6.3f}  max={max(values):6.3f}")
    # Write detailed CSV for further analysis
    if combo_timings_rows:
        import csv
        fieldnames = sorted({k for row in combo_timings_rows for k in row.keys()})
        try:
            write_header = not os.path.exists(TIMING_CSV)
            with open(TIMING_CSV, "a", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fieldnames)
                if write_header:
                    w.writeheader()
                w.writerows(combo_timings_rows)
            print(f"Timing details appended to {TIMING_CSV} ({len(combo_timings_rows)} rows).")
        except Exception as e:
            print(f"[Timing] Failed to write {TIMING_CSV}: {e}")

# --- Global in-memory cache of existing results ---
# Keyed by (Symbol, ATR Multiplier, RR, Vol Multiplier) -> full row dict
existing_rows = {}
AUTOSAVE_ROWS_THRESHOLD = 200  # Autosave every N tests

def _key(symbol, atr, rr, vol):
    try:
        return (str(symbol), float(atr), float(rr), float(vol))
    except Exception:
        return (str(symbol), atr, rr, vol)

# Ensure consistent dtypes & column order
EXPECTED_COLS = [
    "Symbol", "ATR Multiplier", "RR", "Vol Multiplier",
    "Net Profit", "Net Profit Clean", "Win Rate", "drawdown",
    "Total Trades", "Profit Factor"
]


# --- Unified formatting used by both autosave and final save ---
def _format_results_df(results_list, symbol_name):
    df = pd.DataFrame(results_list)
    if df.empty:
        return df
    # Ensure all expected columns exist
    for col in ["ATR Multiplier", "RR", "Vol Multiplier", "Net Profit", "Win Rate", "drawdown", "Total Trades", "Profit Factor"]:
        if col not in df.columns:
            df[col] = pd.NA
    df["Symbol"] = symbol_name
    # Clean/cast (locale-safe: keep decimal commas by converting to dots first)
    df["Win Rate"] = df["Win Rate"].astype(str).str.replace('%', '', regex=False).str.replace(',', '.', regex=False)
    df["drawdown"] = df["drawdown"].astype(str) \
        .str.replace('%', '', regex=False) \
        .str.replace('€', '', regex=False) \
        .str.replace('£', '', regex=False) \
        .str.replace(',', '.', regex=False)
    df["Profit Factor"] = df["Profit Factor"].astype(str).str.replace('x', '', regex=False).str.replace(',', '.', regex=False)
    # remove thin spaces and non-digits
    df["Total Trades"] = df["Total Trades"].astype(str).str.replace('\u202f', '', regex=False).str.replace(' ', '', regex=False).str.replace(r'[^\d]', '', regex=True)

    df["Win Rate"] = pd.to_numeric(df["Win Rate"], errors='coerce')
    df["drawdown"] = pd.to_numeric(df["drawdown"], errors='coerce')
    df["Profit Factor"] = pd.to_numeric(df["Profit Factor"], errors='coerce')
    df["Total Trades"] = pd.to_numeric(df["Total Trades"], errors='coerce')
    df["ATR Multiplier"] = pd.to_numeric(df["ATR Multiplier"], errors='coerce')
    df["RR"] = pd.to_numeric(df["RR"], errors='coerce')
    df["Vol Multiplier"] = pd.to_numeric(df["Vol Multiplier"], errors='coerce')

    # More robust Net Profit cleaning: normalize unicode minus/en dash/non-breaking spaces, keep only digits, sign, and dot after standardizing decimal comma and removing thin spaces/currency codes like JPY/USD
    df["Net Profit Clean"] = pd.to_numeric(
        df["Net Profit"].astype(str)
            .str.replace('\u2212', '-', regex=False)  # normalize unicode minus to ASCII
            .str.replace('\u2013', '-', regex=False)  # normalize en dash to ASCII minus
            .str.replace('\u202f', '', regex=False)   # thin space
            .str.replace('\xa0', '', regex=False)     # non-breaking space
            .str.replace(' ', '', regex=False)
            .str.replace(',', '.', regex=False)       # decimal comma -> dot
            .str.replace(r'[^0-9.\-\+\.]', '', regex=True),  # keep digits, dot, +, -
        errors='coerce'
    )

    # Order columns
    ordered_cols = EXPECTED_COLS
    remaining = [c for c in df.columns if c not in ordered_cols]
    df = df[ordered_cols + remaining]
    return df


# --- Autosave helper: write partial results every N tests (same format as final) ---
from openpyxl import load_workbook

def _compute_best_row(df_symbol: pd.DataFrame):
    if df_symbol.empty:
        return None
    if "Net Profit Clean" not in df_symbol.columns:
        return None
    if df_symbol["Net Profit Clean"].dropna().empty:
        return None
    return df_symbol.loc[df_symbol["Net Profit Clean"].idxmax()]


def _refresh_cache_from_df(df_symbol: pd.DataFrame):
    for _, r in df_symbol.iterrows():
        k = _key(r.get("Symbol"), r.get("ATR Multiplier"), r.get("RR"), r.get("Vol Multiplier"))
        existing_rows[k] = {col: r.get(col) for col in df_symbol.columns}


def autosave_and_update(xlsx_path: str, symbol_name: str, results_list: list):
    if not results_list:
        return
    # Normalize partial into consistent DF
    df_partial = _format_results_df(results_list, symbol_name)

    # Merge with existing symbol sheet (if any) and dedup by key
    if os.path.exists(xlsx_path):
        try:
            book = load_workbook(xlsx_path)
            if f"{symbol_name}_Results" in book.sheetnames:
                df_prev = pd.read_excel(xlsx_path, sheet_name=f"{symbol_name}_Results", engine='openpyxl')
                df_prev = _format_results_df(df_prev.to_dict(orient='records'), symbol_name)
                df_symbol = pd.concat([df_prev, df_partial], ignore_index=True)
            else:
                df_symbol = df_partial.copy()
        except Exception:
            df_symbol = df_partial.copy()
    else:
        df_symbol = df_partial.copy()

    # Prefer rows that have drawdown/Total Trades/Profit Factor filled, then higher profit
    df_symbol["__complete__"] = df_symbol[["drawdown", "Total Trades", "Profit Factor"]].notna().sum(axis=1)
    df_symbol = df_symbol.sort_values(
        ["Symbol", "ATR Multiplier", "RR", "Vol Multiplier", "__complete__", "Net Profit Clean"],
        ascending=[True, True, True, True, False, False]
    ).drop_duplicates(
        subset=["Symbol", "ATR Multiplier", "RR", "Vol Multiplier"],
        keep="first"
    ).drop(columns="__complete__", errors="ignore")

    # Best row for this symbol
    best_row = _compute_best_row(df_symbol)

    # Rebuild Best_Per_Symbol by replacing/adding this symbol
    if os.path.exists(xlsx_path):
        try:
            if "Best_Per_Symbol" in load_workbook(xlsx_path).sheetnames:
                best_df_prev = pd.read_excel(xlsx_path, sheet_name="Best_Per_Symbol", engine='openpyxl')
                best_df_prev = best_df_prev[best_df_prev["Symbol"] != symbol_name]
            else:
                best_df_prev = pd.DataFrame()
        except Exception:
            best_df_prev = pd.DataFrame()
    else:
        best_df_prev = pd.DataFrame()

    if best_row is not None:
        best_df_new = pd.DataFrame({
            "Symbol": [symbol_name],
            "Best ATR Multiplier": [best_row.get("ATR Multiplier")],
            "Best RR": [best_row.get("RR")],
            "Best Vol Multiplier": [best_row.get("Vol Multiplier")],
            "Best Net Profit": [best_row.get("Net Profit")],
            "Best Win Rate": [best_row.get("Win Rate")],
            "Best Drawdown": [best_row.get("drawdown")],
            "Best Total Trades": [best_row.get("Total Trades")],
            "Best Profit Factor": [best_row.get("Profit Factor")],
            "Best Net Profit Clean": [best_row.get("Net Profit Clean")],
        })
        best_df_merged = pd.concat([best_df_prev, best_df_new], ignore_index=True)
    else:
        best_df_merged = best_df_prev

    # Write symbol sheet + Best_Per_Symbol
    mode = 'a' if os.path.exists(xlsx_path) else 'w'
    with pd.ExcelWriter(xlsx_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
        df_symbol.to_excel(writer, index=False, sheet_name=f"{symbol_name}_Results")
        if not best_df_merged.empty:
            best_df_merged.to_excel(writer, index=False, sheet_name="Best_Per_Symbol")

    # Refresh in-memory cache from latest df_symbol
    _refresh_cache_from_df(df_symbol)

    # Rebuild All_Results by concatenating all *_Results sheets
    try:
        book = load_workbook(xlsx_path)
        all_dfs = []
        for sh in book.sheetnames:
            if sh.endswith("_Results"): # Exclude All_results
                if sh == "All_Results":
                    continue
                df_sh = pd.read_excel(xlsx_path, sheet_name=sh, engine='openpyxl')
                df_sh = _format_results_df(df_sh.to_dict(orient='records'), sh.replace("_Results", ""))
                all_dfs.append(df_sh)
        if all_dfs:
            all_merged = pd.concat(all_dfs, ignore_index=True)
            all_merged["__complete__"] = all_merged[["drawdown", "Total Trades", "Profit Factor"]].notna().sum(axis=1)
            all_merged = all_merged.sort_values(
                ["Symbol", "ATR Multiplier", "RR", "Vol Multiplier", "__complete__", "Net Profit Clean"],
                ascending=[True, True, True, True, False, False]
            ).drop_duplicates(
                subset=["Symbol", "ATR Multiplier", "RR", "Vol Multiplier"],
                keep="first"
            ).drop(columns="__complete__", errors="ignore")
            with pd.ExcelWriter(xlsx_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                all_merged.to_excel(writer, index=False, sheet_name="All_Results")
    except Exception as e:
        print(f"[AutoSave] Failed to rebuild All_Results: {e}")

    print(f"[AutoSave] {symbol_name}: {len(df_symbol)} unique rows saved. Best updated.")
    try:
        sys.stdout.write("\n"); sys.stdout.flush()
    except Exception:
        pass

def finalize_symbol(xlsx_path: str, symbol_name: str):
    """Force a final formatting pass on the symbol's sheet so the last batch is properly formatted."""
    try:
        df_raw = pd.read_excel(xlsx_path, sheet_name=f"{symbol_name}_Results", engine='openpyxl')
        df_fmt = _format_results_df(df_raw.to_dict(orient='records'), symbol_name)
        with pd.ExcelWriter(xlsx_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_fmt.to_excel(writer, index=False, sheet_name=f"{symbol_name}_Results")
        print(f"[Finalize] {symbol_name}: sheet formatted and rewritten.")
    except Exception as e:
        print(f"[Finalize] Failed to finalize {symbol_name}: {e}")

# === CONFIG ===
TRADINGVIEW_URL = "https://www.tradingview.com/chart/0RKjg68o/"

# === NIVEAU DE TEST CONFIGURABLE ===
TEST_LEVELS = {
    'COARSE': {
        'ATR_MULTIPLIERS': [1.0, 1.5, 2.0, 2.5, 3.0],  # 5 valeurs
        'RR_VALUES': [2.0, 2.5, 3.0, 3.5, 4.0],        # 5 valeurs  
        'VOL_MULTIPLIERS': [0.8, 1.0, 1.2],            # 3 valeurs
        'description': 'Test rapide avec 75 combinaisons par symbole'
    },
    'FINE': {
        'ATR_MULTIPLIERS': [round(i * 0.2, 1) for i in range(5, 16)],  # 1.0 à 3.0 par 0.2 = 11 valeurs
        'RR_VALUES': [round(i * 0.2, 1) for i in range(10, 21)],       # 2.0 à 4.0 par 0.2 = 11 valeurs
        'VOL_MULTIPLIERS': [round(i * 0.1, 1) for i in range(8, 14)],  # 0.8 à 1.3 par 0.1 = 6 valeurs
        'description': 'Test intermédiaire avec 726 combinaisons par symbole'
    },
    'FULL': {
        'ATR_MULTIPLIERS': [round(i * 0.1, 1) for i in range(10, 31)],  # 1.0 à 3.0 par 0.1 = 21 valeurs
        'RR_VALUES': [round(i * 0.1, 1) for i in range(20, 51)],        # 2.0 à 5.0 par 0.1 = 31 valeurs
        'VOL_MULTIPLIERS': [round(i * 0.1, 1) for i in range(8, 14)],   # 0.8 à 1.3 par 0.1 = 6 valeurs
        'description': 'Test complet avec 3906 combinaisons par symbole'
    }
}

SYMBOL_LIST = ['EURUSD', 'EURAUD', 'USDCAD', 'NZDJPY', 'GBPUSD', 'USDJPY', 'EURJPY', 'GBPJPY', 'AUDUSD', 'AUDJPY', 'AUDCAD', 'USDCHF', 'EURNZD', 'EURGBP', 'NZDUSD', 'EURCAD', 'EURCHF', 'GBPCAD', 'AUDNZD', 'CADCHF', 'GBPCHF', 'CADJPY', 'GBPAUD', 'GBPNZD', 'NZDCAD']

def set_test_level(level='FINE'):
    """Configure le niveau de test et retourne les paramètres correspondants."""
    if level not in TEST_LEVELS:
        print(f"❌ Niveau '{level}' inconnu. Niveaux disponibles: {list(TEST_LEVELS.keys())}")
        level = 'FINE'
    
    config = TEST_LEVELS[level]
    total_combos = len(config['ATR_MULTIPLIERS']) * len(config['RR_VALUES']) * len(config['VOL_MULTIPLIERS'])
    
    print(f"\n📊 NIVEAU DE TEST: {level}")
    print(f"📈 {config['description']}")
    print(f"🔢 ATR: {len(config['ATR_MULTIPLIERS'])} valeurs de {min(config['ATR_MULTIPLIERS'])} à {max(config['ATR_MULTIPLIERS'])}")
    print(f"🔢 RR: {len(config['RR_VALUES'])} valeurs de {min(config['RR_VALUES'])} à {max(config['RR_VALUES'])}")
    print(f"🔢 Vol: {len(config['VOL_MULTIPLIERS'])} valeurs de {min(config['VOL_MULTIPLIERS'])} à {max(config['VOL_MULTIPLIERS'])}")
    print(f"⚡ Total par symbole: {total_combos:,} combinaisons")
    print("")
    
    return config['ATR_MULTIPLIERS'], config['RR_VALUES'], config['VOL_MULTIPLIERS']

# === Setup Chrome Remote Debugging Attach ===
options = Options()
options.add_argument("--window-size=1920,1080")
options.debugger_address = "127.0.0.1:9222"
driver = webdriver.Chrome(options=options)

# --- Argument parser pour options CLI ---
parser = argparse.ArgumentParser(description="Backtest TradingView avec Selenium")
parser.add_argument('--skip-complete', action='store_true', help='Ignorer les devises déjà complètes (tous les combos testés)')
parser.add_argument('--level', choices=['COARSE', 'FINE', 'FULL'], default='FINE', 
                   help='Niveau de test: COARSE (rapide, 75 combos), FINE (moyen, 726 combos), FULL (complet, 3906 combos)')
parser.add_argument('--symbols', nargs='*', help='Symboles spécifiques à tester (ex: --symbols EURUSD GBPUSD)')
args = parser.parse_args()

# Configuration du niveau de test
ATR_MULTIPLIERS, RR_VALUES, VOL_MULTIPLIERS = set_test_level(args.level)

# Filtrer les symboles si spécifié
if args.symbols:
    # Valider que les symboles existent
    invalid_symbols = [s for s in args.symbols if s not in SYMBOL_LIST]
    if invalid_symbols:
        print(f"❌ Symboles invalides: {invalid_symbols}")
        print(f"📋 Symboles disponibles: {SYMBOL_LIST}")
        exit(1)
    SYMBOL_LIST = args.symbols
    print(f"🎯 Test limité à {len(SYMBOL_LIST)} symbole(s): {', '.join(SYMBOL_LIST)}")

# --- ETA totals (calculés après la configuration) ---
TOTAL_COMBOS_PER_SYMBOL = len(ATR_MULTIPLIERS) * len(RR_VALUES) * len(VOL_MULTIPLIERS)
TOTAL_SYMBOLS = len(SYMBOL_LIST)
GLOBAL_TOTAL_COMBOS = TOTAL_COMBOS_PER_SYMBOL * TOTAL_SYMBOLS
global_done_combos = 0

# Affichage des totaux finaux
print(f"🌍 Total pour {TOTAL_SYMBOLS} symbole(s): {GLOBAL_TOTAL_COMBOS:,} tests")
avg_time_per_test = 4.0
estimated_hours = (GLOBAL_TOTAL_COMBOS * avg_time_per_test) / 3600
if estimated_hours < 1:
    print(f"⏱️ Temps estimé total: ~{estimated_hours*60:.0f} minutes")
else:
    print(f"⏱️ Temps estimé total: ~{estimated_hours:.1f}h")
print("")

# === Load existing results to skip already-tested combinations ===
output_file = f"tradingview_backtest_results_{args.level.lower()}.xlsx"
if os.path.exists(output_file):
    try:
        book = load_workbook(output_file)
        for sheet_name in book.sheetnames:
            if not sheet_name.endswith("_Results"):
                continue
            symbol = sheet_name.replace("_Results", "")
            df_prev = pd.read_excel(output_file, sheet_name=sheet_name, engine="openpyxl")
            df_prev = _format_results_df(df_prev.to_dict(orient='records'), symbol)
            for _, row in df_prev.iterrows():
                k = _key(symbol, row.get("ATR Multiplier"), row.get("RR"), row.get("Vol Multiplier"))
                existing_rows[k] = {col: row.get(col) for col in df_prev.columns}
        print(f"Loaded {len(existing_rows)} cached rows from workbook.")
    except Exception as e:
        print(f"Could not load existing results: {e}")
        # Create the xlsx file if it doesn't exist
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, index=False, sheet_name="All_Results")
            pd.DataFrame().to_excel(writer, index=False, sheet_name="Best_Per_Symbol")
else:
    # Initialize a new workbook with expected sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, index=False, sheet_name="All_Results")
        pd.DataFrame().to_excel(writer, index=False, sheet_name="Best_Per_Symbol")

# Navigate to the correct chart
driver.get(TRADINGVIEW_URL)

# Save full HTML for inspection
html = driver.page_source
# Smart wait for the page to load instead of fixed sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

print("Log into TradingView manually if needed. Starting tests...")

# Confirmation pour les tests longs
if args.level == 'FULL':
    estimated_hours = (GLOBAL_TOTAL_COMBOS * 4.0) / 3600
    print(f"⚠️  ATTENTION: Vous avez sélectionné le niveau FULL")
    print(f"⏱️  Temps estimé: ~{estimated_hours:.1f}h ({estimated_hours*24:.1f} jours)")
    confirm = input("🤔 Êtes-vous sûr de vouloir continuer? (tapez 'OUI' pour confirmer): ")
    if confirm != 'OUI':
        print("❌ Test annulé.")
        driver.quit()
        exit()

print(f"\n🚀 DÉMARRAGE DU TEST NIVEAU {args.level}")
print(f"📊 {GLOBAL_TOTAL_COMBOS:,} tests au total")

results = []

# Get all currencies in the list
# currencies = driver.find_elements(By.XPATH, "//div[@data-symbol-full]")
# currency_divs = [{'name': currency.text.split("\n")[0], 'element' : currency} for currency in currencies]
# print("Available currencies:", [currency['name'] for currency in currency_divs])

for currency in SYMBOL_LIST:
    results = []
    tests_since_last_save = 0
    symbol_done = 0
    symbol_total = TOTAL_COMBOS_PER_SYMBOL
    print("")  # ensure a fresh line for progress
    symbol_name = currency
    # Option pour skip les devises déjà complètes (activable via CLI)
    if args.skip_complete:
        combos_tested = 0
        for atr in ATR_MULTIPLIERS:
            for rr in RR_VALUES:
                for vol_mult in VOL_MULTIPLIERS:
                    row_key = _key(symbol_name, atr, rr, vol_mult)
                    if row_key in existing_rows:
                        combos_tested += 1
        if combos_tested >= TOTAL_COMBOS_PER_SYMBOL:
            print(f"[SKIP] {symbol_name}: tous les combos déjà testés ({combos_tested}/{TOTAL_COMBOS_PER_SYMBOL})")
            continue
    # Click on the currency to open its chart
    with timed("symbol_change"):
        driver.get(f"https://www.tradingview.com/chart/0RKjg68o/?symbol=PEPPERSTONE:{currency}")
        time.sleep(5)  # Wait for the chart to load
    print(f"Testing symbol: {symbol_name}")
    # Open settings using Command+P (Mac)
    actions = ActionChains(driver)
    with timed("open_settings_cmdP"):
        actions.key_down(Keys.COMMAND).send_keys('p').key_up(Keys.COMMAND).perform()
    time.sleep(2)
    for atr in ATR_MULTIPLIERS:
        for rr in RR_VALUES:
            for vol_mult in VOL_MULTIPLIERS:
                # === Skip already-tested combinations (use in-memory cache) ===
                row_key = _key(symbol_name, atr, rr, vol_mult)
                if row_key in existing_rows:
                    # Append cached row so we keep a full local results list for symbol
                    cached = existing_rows[row_key].copy()
                    # Ensure Symbol column matches current symbol (sheet merge safety)
                    cached["Symbol"] = symbol_name
                    results.append(cached)
                    tests_since_last_save += 1
                    if tests_since_last_save >= AUTOSAVE_ROWS_THRESHOLD:
                        with timed("autosave"):
                            autosave_and_update(output_file, symbol_name, results)
                        tests_since_last_save = 0
                    # progress for skipped (cached) combo
                    symbol_done += 1
                    global_done_combos += 1
                    _print_progress(symbol_name, symbol_done, symbol_total, global_done_combos, GLOBAL_TOTAL_COMBOS)
                    continue
                success = False
                for attempt in range(3):
                    print(f"Testing ATR={atr}, RR={rr}, VM={vol_mult}, Attempt={attempt+1}")
                    combo_t0 = time.perf_counter()
                    set_inputs_t0 = time.perf_counter()
                    with timed("edit_inputs_total"):
                        # ATR input
                        with timed("edit_input_ATR"):
                            atr_input = driver.find_element(
                                By.XPATH,
                                "//div[contains(text(),'ATR Stop Multiplier') or contains(text(),'Multiplicateur ATR')]/parent::div/following-sibling::div//input"
                            )
                            atr_input.send_keys(Keys.COMMAND + "a")
                            actions.key_down(Keys.COMMAND).send_keys('a').key_up(Keys.COMMAND).perform()
                            time.sleep(0.1)
                            atr_input.send_keys(Keys.BACKSPACE)
                            atr_input.send_keys(str(atr))

                        # RR input
                        with timed("edit_input_RR"):
                            rr_input = driver.find_element(
                                By.XPATH,
                                "//div[contains(text(),'Base Risk/Reward Ratio') or contains(text(),'Ratio Risque/Rendement de base')]/parent::div/following-sibling::div//input"
                            )
                            rr_input.send_keys(Keys.COMMAND + "a")
                            actions.key_down(Keys.COMMAND).send_keys('a').key_up(Keys.COMMAND).perform()
                            rr_input.send_keys(Keys.BACKSPACE)
                            rr_input.send_keys(str(rr))

                        # Volatility Multiplier input (Dynamic: Min Volatility Multiplier)
                        with timed("edit_input_VolMult"):
                            vol_input = driver.find_element(
                                By.XPATH,
                                "//div[contains(text(),'Dynamic: Min Volatility Multiplier') or contains(text(),'Multiplicateur minimal de volatilité dynamique')]/parent::div/following-sibling::div//input"
                            )
                            vol_input.send_keys(Keys.COMMAND + "a")
                            actions.key_down(Keys.COMMAND).send_keys('a').key_up(Keys.COMMAND).perform()
                            vol_input.send_keys(Keys.BACKSPACE)
                            vol_input.send_keys(str(vol_mult))

                        actions.send_keys(Keys.TAB).perform()
                    with timed("wait_snackbar_hide"):
                        try:
                            WebDriverWait(driver, 3).until_not(
                                EC.presence_of_element_located((By.CLASS_NAME, "snackbarLayer-_MKqWk5g"))
                            )
                        except:
                            pass
                    # Attendre que le texte "Profit net" ou "Total P&L" soit présent
                    with timed("wait_profit_locator"):
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Profit net') or contains(text(),'Total P&L')]"))
                        )
                    with timed("scrape_metrics"):
                        # Get Net Profit
                        net_profit_elem = driver.find_element(By.XPATH, "//div[contains(text(),'Profit net') or contains(text(),'Total P&L')]/parent::div/following-sibling::div/div[3]")
                        net_profit = net_profit_elem.text
                        # Normalize various unicode minus/dash characters and spaces so the sign is preserved
                        net_profit = net_profit.replace('\u2212', '-')  # unicode minus → ASCII minus
                        net_profit = net_profit.replace('\u2013', '-')  # en dash → ASCII minus
                        net_profit = net_profit.replace('\u202f', '')   # thin space
                        net_profit = net_profit.replace('\xa0', '')     # non-breaking space
                        net_profit = net_profit.replace(' ', '')
                        net_profit = net_profit.replace(',', '.')       # decimal comma to dot
                        # Keep only digits, dot, plus and minus (sign)
                        net_profit = re.sub(r'[^0-9.\-\+]', '', net_profit).strip()

                        # Win Rate
                        win_rate_elem = driver.find_element(By.XPATH, "//div[contains(text(),'Pourcentage de trades gagnants') or contains(text(),'Profitable trades')]/parent::div/following-sibling::div/div[1]")
                        win_rate = win_rate_elem.text

                        # Drawdown
                        drawdown_elem = driver.find_element(By.XPATH, "//div[contains(text(),'Drawdown') or contains(text(),'drawdown')]/parent::div/following-sibling::div/div[3]")
                        drawdown = drawdown_elem.text
                        drawdown = drawdown.replace('€', '').replace('£', '').replace('%', '')
                        drawdown = drawdown.replace(',', '.').strip()

                        # Total trades
                        total_trades_elem = driver.find_element(By.XPATH, "//div[contains(text(),'Total des trades') or contains(text(),'Total trades')]/parent::div/following-sibling::div")
                        total_trades = total_trades_elem.text
                        total_trades = total_trades.replace('\u202f', '').replace(' ', '')
                        total_trades = ''.join(ch for ch in total_trades if ch.isdigit())

                        # Profit factor
                        profit_factor_elem = driver.find_element(By.XPATH, "//div[contains(text(),'Profit factor') or contains(text(),'Profit factor')]/parent::div/following-sibling::div")
                        profit_factor = profit_factor_elem.text.replace('x', '').replace(',', '.').strip()

                    print(f"Net Profit: {net_profit}, Win Rate: {win_rate}, Drawdown: {drawdown}, Total Trades: {total_trades}, Profit Factor: {profit_factor}")

                    results.append({
                        "ATR Multiplier": atr,
                        "RR": rr,
                        "Vol Multiplier": vol_mult,
                        "Net Profit": net_profit,
                        "Win Rate": win_rate,
                        "drawdown": drawdown,
                        "Total Trades": total_trades,
                        "Profit Factor": profit_factor
                    })
                    combo_dt = time.perf_counter() - combo_t0
                    _append_combo_timing_row(symbol_name, atr, rr, vol_mult,
                                             combo_total=combo_dt)
                    tests_since_last_save += 1
                    if tests_since_last_save >= AUTOSAVE_ROWS_THRESHOLD:
                        with timed("autosave"):
                            autosave_and_update(output_file, symbol_name, results)
                        tests_since_last_save = 0
                    combo_time_window.append(combo_dt)
                    symbol_done += 1
                    global_done_combos += 1
                    _print_progress(symbol_name, symbol_done, symbol_total, global_done_combos, GLOBAL_TOTAL_COMBOS)
                    success = True
                    break
                if not success:
                    results.append({
                        "ATR Multiplier": atr,
                        "RR": rr,
                        "Vol Multiplier": vol_mult,
                        "Net Profit": "Error",
                        "Win Rate": "Error",
                        "drawdown": "Error",
                        "Total Trades": "Error",
                        "Profit Factor": "Error"
                    })
                    tests_since_last_save += 1
                    if tests_since_last_save >= AUTOSAVE_ROWS_THRESHOLD:
                        with timed("autosave"):
                            autosave_and_update(output_file, symbol_name, results)
                        tests_since_last_save = 0
                    combo_time_window.append(_avg_combo_seconds())
                    symbol_done += 1
                    global_done_combos += 1
                    _print_progress(symbol_name, symbol_done, symbol_total, global_done_combos, GLOBAL_TOTAL_COMBOS)
                time.sleep(0.1)

    # Final autosave for any remaining unsaved results for this symbol
    with timed("autosave"):
        autosave_and_update(output_file, symbol_name, results)
    print("")  # finalize the progress line for this symbol
    # Results are already persisted & deduplicated by autosave_and_update.
    # Proceed with parameter reset for the next symbol.
    with timed("finalize_symbol"):
        finalize_symbol(output_file, symbol_name)
    # Reset to base parameters
    with timed("reset_params"):
        BASE_ATR = 1.2
        BASE_RR = 2.7
        BASE_VOL_MULT = 0.8
        actions = ActionChains(driver)
        actions.key_down(Keys.COMMAND).send_keys('p').key_up(Keys.COMMAND).perform()
        time.sleep(2)

        # ATR input
        atr_input = driver.find_element(
            By.XPATH,
            "//div[contains(text(),'ATR Stop Multiplier') or contains(text(),'Multiplicateur ATR')]/parent::div/following-sibling::div//input"
        )
        atr_input.send_keys(Keys.COMMAND + "a")
        actions.key_down(Keys.COMMAND).send_keys('a').key_up(Keys.COMMAND).perform()
        time.sleep(1)  # Ensure input is cleared
        atr_input.send_keys(Keys.BACKSPACE)
        atr_input.send_keys(str(BASE_ATR))

        # RR input
        rr_input = driver.find_element(
            By.XPATH,
            "//div[contains(text(),'Base Risk/Reward Ratio') or contains(text(),'Ratio Risque/Rendement de base')]/parent::div/following-sibling::div//input"
        )
        rr_input.send_keys(Keys.COMMAND + "a")
        actions.key_down(Keys.COMMAND).send_keys('a').key_up(Keys.COMMAND).perform()
        time.sleep(1)  # Ensure input is cleared
        rr_input.send_keys(Keys.BACKSPACE)
        rr_input.send_keys(str(BASE_RR))

        # Volatility Multiplier input (Dynamic: Min Volatility Multiplier)
        vol_input = driver.find_element(
            By.XPATH,
            "//div[contains(text(),'Dynamic: Min Volatility Multiplier') or contains(text(),'Multiplicateur minimal de volatilité dynamique')]/parent::div/following-sibling::div//input"
        )
        vol_input.send_keys(Keys.COMMAND + "a")
        actions.key_down(Keys.COMMAND).send_keys('a').key_up(Keys.COMMAND).perform()
        vol_input.send_keys(Keys.BACKSPACE)
        vol_input.send_keys(str(BASE_VOL_MULT))

        # Wait until snackbar disappears if present
        try:
            WebDriverWait(driver, 5).until_not(
                EC.presence_of_element_located((By.CLASS_NAME, "snackbarLayer-_MKqWk5g"))
            )
        except:
            pass  # Ignore if not found or timeout
        close_button = driver.find_element(By.XPATH, "//button[@data-name='submit-button']")
        close_button.click()
        time.sleep(2)
        print("Strategy parameters reset to default values.")


# Results are already persisted & deduplicated by autosave_and_update.
# Remove the final manual All_Results rebuild block entirely.
print(f"Backtesting complete. Results saved to {output_file}")

dump_timing_summary()

driver.quit()